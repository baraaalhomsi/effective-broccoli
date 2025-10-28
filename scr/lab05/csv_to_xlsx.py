import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:

    if not Path(csv_path).exists():
        raise FileNotFoundError(f"file {csv_path} is not exist")

    try:
        with open(csv_path, 'r', encoding='utf-8') as csv_file:
            reader = csv.reader(csv_file)
            rows = list(reader)
    except Exception as e:
        raise ValueError(f"wrong while reading CSV: {e}")

    if not rows:
        raise ValueError("file CSV is empty")
    
    # to creat the fail if it is not not exist
    Path(xlsx_path).parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # copy info from CSV to Exel
    for row in rows:
        ws.append(row)
    
    # adjusting width of the colmuns
    for col_num, column in enumerate(ws.columns, 1):
        max_length = 0
        
        # the longest text in te column
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # adjusting width
        adjusted_width = max(8, min(max_length + 2, 50))
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(xlsx_path)

def main():

    try:
        print("converting people.csv to XLSX...")
        csv_to_xlsx("data/samples/people.csv", "data/out/people.xlsx")
        print("the conversion was successful: people.xlsx")
        
        print("converting cities.csv to XLSX...")
        csv_to_xlsx("data/samples/cities.csv", "data/out/cities.xlsx")
        print("the conversion was successful: cities.xlsx")
        
    except Exception as e:
        print(f"wrong: {e}")


if __name__ == "__main__":
    main()